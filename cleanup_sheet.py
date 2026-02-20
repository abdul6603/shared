"""Clean up Brotherhood Progress Sheet — remove bad rows, fix descriptions, re-sequence."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

SRC = Path.home() / "Desktop" / "brotherhood_progress.xlsx"
DEST_PATHS = [
    Path.home() / "Desktop" / "brotherhood_progress.xlsx",
    Path.home() / "thor" / "data" / "brotherhood_progress.xlsx",
]

# === Rows to DELETE (by original seq number) ===
DELETE_SEQS = {
    72,   # Competitor Content Scraper — MISSING (no code)
    185,  # Robotox Singleton Pattern — MISSING (never implemented)
    301,  # Intelligence Feed Publisher — NOT BUILT
    304,  # Intelligence Feed Consumer — NOT BUILT
    309,  # Intel Feed Dashboard Panel — NOT BUILT
}

# === Rows to FIX (seq -> new values) ===
# Format: seq -> dict of column overrides (change, description, type, agent)
FIX_ROWS = {
    16: {"change": "ElevenLabs Voice (Titan)",
         "description": "Voice integration using ElevenLabs API with Titan voice (voice_id: f5KRUAmxOzuhrrp8V3zv). OpenAI TTS as fallback. Full voiceover pipeline in generate.py."},
    23: {"change": "Flask Control Server",
         "description": "Flask server on port 7777 with blueprint routes from core/routes.py. Scheduler, economics, task management, broadcast, decision memory APIs."},
    60: {"change": "GitHub Commit Capability",
         "description": "Git commit/push capability described in brotherhood_commands.json and knowledge base. Execution handled by Claude or manual — no autonomous git code in Thor."},
    66: {"change": "Flask Backend",
         "description": "Flask + SocketIO backend with 14 route blueprints. Real-time WebSocket heartbeats, health push events. LaunchAgent auto-restart on deploy."},
    125: {"change": "Indicator Accuracy Deprioritized",
         "description": "Indicator accuracy table removed from overview dashboard UI. Backend code still exists in shared.py, garves routes, tracker, weight_learner for internal use."},
    187: {"change": "Process Monitor + LaunchAgent Config",
         "description": "Process monitoring via pattern matching in Robotox monitor.py. LaunchAgent labels defined in config for Hawk/Viper but launchctl management handled by dashboard system routes, not Robotox directly."},
}

# === Theme colors (same as redesign_sheet.py) ===
HEADER_BG = "1A1A2E"
HEADER_FG = "FFFFFF"
HEADER_ACCENT = "FFD700"
BORDER_COLOR = "2A2A3E"

TYPE_COLORS = {
    "Feature":     {"badge_bg": "1E3A5F", "badge_fg": "5DADE2", "row_bg": "0D1B2A"},
    "Fix":         {"badge_bg": "5C2D00", "badge_fg": "FF9F43", "row_bg": "1A1200"},
    "Upgrade":     {"badge_bg": "0D3B1E", "badge_fg": "2ECC71", "row_bg": "0A1A0D"},
    "Integration": {"badge_bg": "2D1B4E", "badge_fg": "A569BD", "row_bg": "150D22"},
}
DEFAULT_TC = {"badge_bg": "2A2A2A", "badge_fg": "AAAAAA", "row_bg": "111111"}

AGENT_COLORS = {
    "Garves": "00D4FF", "Soren": "CC66FF", "Shelby": "FFAA00",
    "Atlas": "22AA44", "Lisa": "FF8800", "Robotox": "00FF44",
    "Thor": "FF6600", "Hawk": "FFD700", "Viper": "00FF88",
    "System": "888888", "Dashboard": "888888", "Lisa+Soren": "CC66FF",
}

COLUMNS = [
    ("#",           6,   "center"),
    ("Date",        16,  "center"),
    ("Time",        12,  "center"),
    ("Agent",       14,  "center"),
    ("Type",        14,  "center"),
    ("Change",      40,  "left"),
    ("Description", 80,  "left"),
    ("Duration",    14,  "center"),
    ("Status",      12,  "center"),
]

thin_side = Side(style="thin", color=BORDER_COLOR)
THIN_BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
DARK_BG = PatternFill(start_color="0F0F1A", end_color="0F0F1A", fill_type="solid")


def style_row(ws, row_num, agent, type_val):
    tc = TYPE_COLORS.get(type_val, DEFAULT_TC)
    row_fill = PatternFill(start_color=tc["row_bg"], end_color=tc["row_bg"], fill_type="solid")
    badge_fill = PatternFill(start_color=tc["badge_bg"], end_color=tc["badge_bg"], fill_type="solid")
    agent_color = AGENT_COLORS.get(agent, "AAAAAA")
    col_align = ["center", "center", "center", "center", "center", "left", "left", "center", "center"]

    ws.row_dimensions[row_num].height = 28

    for col in range(1, 10):
        cell = ws.cell(row=row_num, column=col)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal=col_align[col-1], vertical="center", wrap_text=(col in (6, 7)))

        if col == 1:
            cell.font = Font(name="Calibri", size=12, color="666688")
            cell.fill = row_fill
        elif col == 4:
            cell.font = Font(name="Calibri", size=13, bold=True, color=agent_color)
            cell.fill = row_fill
        elif col == 5:
            cell.font = Font(name="Calibri", size=12, bold=True, color=tc["badge_fg"])
            cell.fill = badge_fill
        elif col == 6:
            cell.font = Font(name="Calibri", size=13, bold=True, color="E0E0E0")
            cell.fill = row_fill
        elif col == 7:
            cell.font = Font(name="Calibri", size=12, color="B0B0C0")
            cell.fill = row_fill
        elif col == 9:
            status_val = str(cell.value or "")
            if status_val == "Done":
                cell.font = Font(name="Calibri", size=12, bold=True, color="2ECC71")
            elif status_val == "In Progress":
                cell.font = Font(name="Calibri", size=12, bold=True, color="F39C12")
            else:
                cell.font = Font(name="Calibri", size=12, color="AAAAAA")
            cell.fill = row_fill
        else:
            cell.font = Font(name="Calibri", size=12, color="9999AA")
            cell.fill = row_fill


def clean():
    wb_old = openpyxl.load_workbook(str(SRC))
    ws_old = wb_old.active

    # Read all rows
    all_rows = []
    for row in ws_old.iter_rows(min_row=2, values_only=True):
        all_rows.append(list(row))

    print(f"Read {len(all_rows)} total rows (including empty)")

    # Filter: keep only valid rows
    clean_rows = []
    removed = {"empty": 0, "deleted": 0, "malformed": 0, "fixed": 0}

    for row in all_rows:
        seq = row[0]
        agent = str(row[3] or "") if len(row) > 3 else ""
        type_val = str(row[4] or "") if len(row) > 4 else ""
        change = str(row[5] or "") if len(row) > 5 else ""

        # Skip empty/null rows
        if seq is None or agent == "" or agent == "None":
            removed["empty"] += 1
            continue

        # Skip malformed rows (date leaked into seq)
        if isinstance(seq, str) and seq.startswith("2026"):
            removed["malformed"] += 1
            continue

        # Try to get numeric seq
        try:
            seq_num = int(seq)
        except (ValueError, TypeError):
            removed["malformed"] += 1
            continue

        # Skip rows marked for deletion
        if seq_num in DELETE_SEQS:
            removed["deleted"] += 1
            print(f"  DELETED #{seq_num}: {change[:50]}")
            continue

        # Fix agent names that are actually descriptions
        if len(agent) > 20:
            agent = "System"
            row[3] = "System"

        # Apply fixes
        if seq_num in FIX_ROWS:
            fixes = FIX_ROWS[seq_num]
            if "change" in fixes:
                row[5] = fixes["change"]
            if "description" in fixes:
                row[6] = fixes["description"]
            if "type" in fixes:
                row[4] = fixes["type"]
            if "agent" in fixes:
                row[3] = fixes["agent"]
            removed["fixed"] += 1
            print(f"  FIXED #{seq_num}: {fixes.get('change', change[:50])}")

        clean_rows.append(row)

    print(f"\nCleanup: {removed['empty']} empty, {removed['malformed']} malformed, "
          f"{removed['deleted']} deleted, {removed['fixed']} fixed")
    print(f"Clean rows: {len(clean_rows)}")

    # Build new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Brotherhood Progress"
    ws.sheet_properties.tabColor = "FFD700"

    # Set column widths
    for i, (header, width, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header row
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(name="Calibri", size=16, bold=True, color=HEADER_FG)
    header_border = Border(
        left=thin_side, right=thin_side,
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="medium", color=HEADER_ACCENT),
    )

    for i, (header, _, align) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # Write clean rows with re-sequenced numbers
    for idx, row in enumerate(clean_rows):
        row_num = idx + 2
        new_seq = idx + 1

        # Pad row to 9 columns if needed (old 8-col rows)
        while len(row) < 9:
            row.append(None)

        date_val = row[1] or ""
        time_val = row[2] or ""
        agent = str(row[3] or "System")
        type_val = str(row[4] or "Feature")
        change = str(row[5] or "")
        desc = str(row[6] or "")
        duration = str(row[7] or "--") if row[7] and str(row[7]) != "None" else "--"

        # Old format had status in col 8 (index 7), new format has duration in 8, status in 9
        # Detect: if col 8 is "Done"/"In Progress"/"Blocked", it's the old status column
        if duration in ("Done", "In Progress", "Blocked"):
            status = duration
            duration = "--"
        else:
            status = str(row[8] or "Done") if row[8] and str(row[8]) != "None" else "Done"

        values = [new_seq, date_val, time_val, agent, type_val, change, desc, duration, status]
        for col_idx, val in enumerate(values):
            ws.cell(row=row_num, column=col_idx + 1, value=val)

        style_row(ws, row_num, agent, type_val)

    # Auto-filter
    last_row = len(clean_rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"

    # Save
    for dest in DEST_PATHS:
        try:
            dest.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(dest))
            print(f"Saved: {dest}")
        except Exception as e:
            print(f"Failed: {dest}: {e}")

    print(f"\nDone! {len(clean_rows)} clean rows written.")


if __name__ == "__main__":
    clean()
