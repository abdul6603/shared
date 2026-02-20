"""One-time redesign of Brotherhood Progress Excel sheet.

New format:
- Bigger fonts (14pt data, 18pt headers)
- Color-coded TYPE column (Feature=blue, Fix=orange, Upgrade=green, Integration=purple)
- Agent names in brand colors
- New DURATION column
- Clean alternating row colors per type
- Frozen header row
- Auto-filter
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path
from copy import copy

SRC = Path.home() / "Desktop" / "brotherhood_progress.xlsx"
DEST_PATHS = [
    Path.home() / "Desktop" / "brotherhood_progress.xlsx",
    Path.home() / "thor" / "data" / "brotherhood_progress.xlsx",
]

# === Color scheme ===
HEADER_BG = "1A1A2E"       # Deep navy
HEADER_FG = "FFFFFF"        # White text
HEADER_ACCENT = "FFD700"    # Gold underline

# Type colors (background tint for the Type cell + subtle row tint)
TYPE_COLORS = {
    "Feature":     {"badge_bg": "1E3A5F", "badge_fg": "5DADE2", "row_bg": "0D1B2A"},  # Blue
    "Fix":         {"badge_bg": "5C2D00", "badge_fg": "FF9F43", "row_bg": "1A1200"},  # Orange
    "Upgrade":     {"badge_bg": "0D3B1E", "badge_fg": "2ECC71", "row_bg": "0A1A0D"},  # Green
    "Integration": {"badge_bg": "2D1B4E", "badge_fg": "A569BD", "row_bg": "150D22"},  # Purple
}
DEFAULT_TYPE_COLOR = {"badge_bg": "2A2A2A", "badge_fg": "AAAAAA", "row_bg": "111111"}

# Agent brand colors
AGENT_COLORS = {
    "Garves":  "00D4FF",
    "Soren":   "CC66FF",
    "Shelby":  "FFAA00",
    "Atlas":   "22AA44",
    "Lisa":    "FF8800",
    "Robotox": "00FF44",
    "Thor":    "FF6600",
    "Hawk":    "FFD700",
    "Viper":   "00FF88",
    "System":  "888888",
    "Dashboard": "888888",
}

# Clean up messy agent names from old data
AGENT_CLEANUP = {}

# Column config: (header, width, align)
COLUMNS = [
    ("#",           6,   "center"),
    ("Date",        16,  "center"),
    ("Time",        12,  "center"),
    ("Agent",       14,  "center"),
    ("Type",        14,  "center"),
    ("Change",      40,  "left"),
    ("Description", 80,  "left"),
    ("Duration",    14,  "center"),
    ("Status",      12,  "center"),
]

THIN_BORDER_COLOR = "2A2A3E"
thin_side = Side(style="thin", color=THIN_BORDER_COLOR)
THIN_BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Dark background for the whole sheet
DARK_BG = PatternFill(start_color="0F0F1A", end_color="0F0F1A", fill_type="solid")


def build_sheet():
    # 1. Read existing data
    wb_old = openpyxl.load_workbook(str(SRC))
    ws_old = wb_old.active
    old_rows = []
    for row in ws_old.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        old_rows.append(list(row))
    print(f"Read {len(old_rows)} existing rows")

    # 2. Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Brotherhood Progress"
    ws.sheet_properties.tabColor = "FFD700"

    # Dark background fill for a huge range
    for r in range(1, len(old_rows) + 100):
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(row=r, column=c).fill = DARK_BG

    # 3. Set column widths
    for i, (header, width, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 4. Header row
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(name="Calibri", size=16, bold=True, color=HEADER_FG)
    header_border = Border(
        left=thin_side, right=thin_side,
        top=Side(style="thin", color=THIN_BORDER_COLOR),
        bottom=Side(style="medium", color=HEADER_ACCENT),
    )

    for i, (header, _, align) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # 5. Data rows
    for idx, old_row in enumerate(old_rows):
        row_num = idx + 2

        # Extract old values (8 columns → 9 columns now)
        seq = old_row[0] if len(old_row) > 0 else idx + 1
        date_val = old_row[1] if len(old_row) > 1 else ""
        time_val = old_row[2] if len(old_row) > 2 else ""
        agent = str(old_row[3] or "") if len(old_row) > 3 else ""
        type_val = str(old_row[4] or "") if len(old_row) > 4 else ""
        change = str(old_row[5] or "") if len(old_row) > 5 else ""
        desc = str(old_row[6] or "") if len(old_row) > 6 else ""
        status = str(old_row[7] or "Done") if len(old_row) > 7 else "Done"

        # Clean up agent names that are actually descriptions
        if len(agent) > 20:
            agent = "System"

        # Fix type that ended up as "Done" (status leaked into type column)
        if type_val == "Done":
            type_val = "Feature"
            # The old status column might be empty, so status = "Done"

        # Duration = "--" for old entries
        duration = "--"

        # Get type colors
        tc = TYPE_COLORS.get(type_val, DEFAULT_TYPE_COLOR)
        row_fill = PatternFill(start_color=tc["row_bg"], end_color=tc["row_bg"], fill_type="solid")
        badge_fill = PatternFill(start_color=tc["badge_bg"], end_color=tc["badge_bg"], fill_type="solid")
        badge_font_color = tc["badge_fg"]

        # Agent color
        agent_color = AGENT_COLORS.get(agent, "AAAAAA")

        ws.row_dimensions[row_num].height = 28

        # Write cells
        values = [seq, date_val, time_val, agent, type_val, change, desc, duration, status]
        for col_idx, val in enumerate(values):
            c = col_idx + 1
            cell = ws.cell(row=row_num, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal=COLUMNS[col_idx][2],
                vertical="center",
                wrap_text=(c in (6, 7)),
            )

            if c == 4:  # Agent — brand color, bold
                cell.font = Font(name="Calibri", size=13, bold=True, color=agent_color)
                cell.fill = row_fill
            elif c == 5:  # Type — badge style
                cell.font = Font(name="Calibri", size=12, bold=True, color=badge_font_color)
                cell.fill = badge_fill
            elif c == 9:  # Status
                if status == "Done":
                    cell.font = Font(name="Calibri", size=12, bold=True, color="2ECC71")
                elif status == "In Progress":
                    cell.font = Font(name="Calibri", size=12, bold=True, color="F39C12")
                else:
                    cell.font = Font(name="Calibri", size=12, color="AAAAAA")
                cell.fill = row_fill
            elif c == 1:  # Row number
                cell.font = Font(name="Calibri", size=12, color="666688")
                cell.fill = row_fill
            elif c == 6:  # Change title — slightly brighter
                cell.font = Font(name="Calibri", size=13, bold=True, color="E0E0E0")
                cell.fill = row_fill
            elif c == 7:  # Description
                cell.font = Font(name="Calibri", size=12, color="B0B0C0")
                cell.fill = row_fill
            else:
                cell.font = Font(name="Calibri", size=12, color="9999AA")
                cell.fill = row_fill

    # 6. Auto-filter
    last_row = len(old_rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"

    # 7. Save
    for dest in DEST_PATHS:
        try:
            dest.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(dest))
            print(f"Saved: {dest}")
        except Exception as e:
            print(f"Failed to save {dest}: {e}")

    print("Done! Sheet redesigned.")


if __name__ == "__main__":
    build_sheet()
