"""Shared progress sheet writer — single source of truth for Excel format.

All writers (Claude, Thor, git hook) should use this to avoid format drift.

Columns (9): # | Date | Time | Agent | Type | Change | Description | Duration | Status

Dark theme with color-coded types and agent brand colors.
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

EXCEL_PATHS = [
    Path.home() / "thor" / "data" / "brotherhood_progress.xlsx",
    Path.home() / "Desktop" / "brotherhood_progress.xlsx",
]

TYPE_MAP = {
    "Feature": "Feature", "New Feature": "Feature",
    "Fix": "Fix", "Bug Fix": "Fix",
    "Upgrade": "Upgrade", "Improvement": "Upgrade",
    "Integration": "Integration",
    "Infrastructure": "Feature",
    "Dashboard": "Upgrade",
}

# Type → badge background, badge text color, row tint
TYPE_COLORS = {
    "Feature":     {"badge_bg": "1E3A5F", "badge_fg": "5DADE2", "row_bg": "0D1B2A"},
    "Fix":         {"badge_bg": "5C2D00", "badge_fg": "FF9F43", "row_bg": "1A1200"},
    "Upgrade":     {"badge_bg": "0D3B1E", "badge_fg": "2ECC71", "row_bg": "0A1A0D"},
    "Integration": {"badge_bg": "2D1B4E", "badge_fg": "A569BD", "row_bg": "150D22"},
}
DEFAULT_TYPE_COLOR = {"badge_bg": "2A2A2A", "badge_fg": "AAAAAA", "row_bg": "111111"}

# Agent brand colors (hex without #)
AGENT_BRAND_COLORS = {
    "Garves": "00D4FF", "Soren": "CC66FF", "Shelby": "FFAA00",
    "Atlas": "22AA44", "Lisa": "FF8800", "Robotox": "00FF44",
    "Thor": "FF6600", "Hawk": "FFD700", "Viper": "00FF88",
    "System": "888888", "Dashboard": "888888",
}

BORDER_COLOR = "2A2A3E"


def _apply_row_style(ws, row_num: int, agent: str, change_type: str) -> None:
    """Apply dark-theme styling to a data row (9 columns)."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return

    thin_side = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    tc = TYPE_COLORS.get(change_type, DEFAULT_TYPE_COLOR)
    row_fill = PatternFill(start_color=tc["row_bg"], end_color=tc["row_bg"], fill_type="solid")
    badge_fill = PatternFill(start_color=tc["badge_bg"], end_color=tc["badge_bg"], fill_type="solid")
    agent_color = AGENT_BRAND_COLORS.get(agent, "AAAAAA")

    # Column alignment: center for #, date, time, agent, type, duration, status; left for change, desc
    col_align = ["center", "center", "center", "center", "center", "left", "left", "center", "center"]

    ws.row_dimensions[row_num].height = 28

    for col in range(1, 10):
        cell = ws.cell(row=row_num, column=col)
        cell.border = border
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=(col in (6, 7)),
            horizontal=col_align[col - 1],
        )

        if col == 1:  # #
            cell.font = Font(name="Calibri", size=12, color="666688")
            cell.fill = row_fill
        elif col == 4:  # Agent — brand color
            cell.font = Font(name="Calibri", size=13, bold=True, color=agent_color)
            cell.fill = row_fill
        elif col == 5:  # Type — badge style
            cell.font = Font(name="Calibri", size=12, bold=True, color=tc["badge_fg"])
            cell.fill = badge_fill
        elif col == 6:  # Change title
            cell.font = Font(name="Calibri", size=13, bold=True, color="E0E0E0")
            cell.fill = row_fill
        elif col == 7:  # Description
            cell.font = Font(name="Calibri", size=12, color="B0B0C0")
            cell.fill = row_fill
        elif col == 9:  # Status
            status_val = str(cell.value or "")
            if status_val == "Done":
                cell.font = Font(name="Calibri", size=12, bold=True, color="2ECC71")
            elif status_val == "In Progress":
                cell.font = Font(name="Calibri", size=12, bold=True, color="F39C12")
            else:
                cell.font = Font(name="Calibri", size=12, color="AAAAAA")
            cell.fill = row_fill
        else:  # Date, Time, Duration
            cell.font = Font(name="Calibri", size=12, color="9999AA")
            cell.fill = row_fill


def append_progress(
    agent: str,
    change_type: str,
    feature: str,
    description: str,
    status: str = "Done",
    duration: str = "--",
) -> int:
    """Append a row to the Brotherhood Progress Excel sheet.

    Args:
        agent: Agent name (Atlas, Thor, Garves, Hawk, etc.)
        change_type: One of Feature, Fix, Upgrade, Integration
        feature: Short title of the change (max 50 chars)
        description: Detailed description (max 200 chars)
        status: Done, In Progress, Blocked
        duration: How long the task took (e.g. "15 min", "2 hrs")

    Returns:
        Number of files successfully written to.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 0

    resolved_type = TYPE_MAP.get(change_type, "Upgrade")

    # Clean description
    desc = re.sub(r'#+\s+', '', description)
    desc = re.sub(r'\*{1,2}([^*]+)\*{1,2}', r'\1', desc)
    desc = re.sub(r'```[\s\S]*?```', '', desc)
    desc = desc.replace('`', '')
    desc = re.sub(r'\s+', ' ', desc).strip()[:200]

    now = datetime.now()
    written = 0

    for excel_path in EXCEL_PATHS:
        try:
            if not excel_path.exists():
                continue
            wb = load_workbook(str(excel_path))
            ws = wb.active

            row_num = ws.max_row + 1
            seq = row_num - 1

            ws.cell(row=row_num, column=1, value=seq)
            ws.cell(row=row_num, column=2, value=now.strftime("%b %d, %Y"))
            ws.cell(row=row_num, column=3, value=now.strftime("%-I:%M %p"))
            ws.cell(row=row_num, column=4, value=agent)
            ws.cell(row=row_num, column=5, value=resolved_type)
            ws.cell(row=row_num, column=6, value=feature[:50])
            ws.cell(row=row_num, column=7, value=desc)
            ws.cell(row=row_num, column=8, value=duration)
            ws.cell(row=row_num, column=9, value=status)

            _apply_row_style(ws, row_num, agent, resolved_type)
            ws.auto_filter.ref = f"A1:I{row_num}"

            wb.save(str(excel_path))
            written += 1
        except Exception:
            continue

    return written
